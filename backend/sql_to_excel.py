import sqlite3
import openpyxl
from pathlib import Path

def export_sqlite_to_excel(db_path, output_file='output.xlsx'):
    """
    Экспортирует все таблицы из SQLite базы данных в один Excel файл,
    где каждая таблица находится на отдельном листе.
    
    :param db_path: Путь к файлу SQLite базы данных
    :param output_file: Имя выходного Excel файла
    """
    # Создаем новый Excel-файл
    wb = openpyxl.Workbook()
    
    # Подключаемся к базе данных
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Получаем список всех таблиц в базе данных
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
    tables = cursor.fetchall()
    
    # Удаляем стандартный лист, если он есть
    if 'Sheet' in wb.sheetnames:
        std_sheet = wb['Sheet']
        wb.remove(std_sheet)
    
    for table in tables:
        table_name = table[0]
        
        # Создаем новый лист с именем таблицы
        ws = wb.create_sheet(title=table_name)
        
        # Получаем данные из таблицы
        cursor.execute(f"SELECT * FROM {table_name};")
        data = cursor.fetchall()
        
        # Получаем названия столбцов
        cursor.execute(f"PRAGMA table_info({table_name});")
        columns = [column[1] for column in cursor.fetchall()]
        
        # Записываем заголовки в первую строку
        for col_num, column in enumerate(columns, 1):
            ws.cell(row=1, column=col_num, value=column)
        
        # Записываем данные
        for row_num, row_data in enumerate(data, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=cell_value)
        
        print(f"Таблица '{table_name}' добавлена на лист '{table_name}'")
    
    # Сохраняем Excel-файл
    wb.save(output_file)
    
    # Закрываем соединение
    conn.close()
    print(f"\nВсе таблицы экспортированы в файл: {output_file}")

# Пример использования
if __name__ == "__main__":
    db_path = './db/pizza.db'
    output_file = './output/output.xlsx'
    if not output_file:
        output_file = 'output.xlsx'
    
    export_sqlite_to_excel(db_path, output_file)